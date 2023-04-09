import platform
import re
import sys
from datetime import datetime as dt
from pathlib import Path

import fire
import openpyxl
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet


def ReadConditionCSV(file_path: Path):
    with open(file_path, mode="r", encoding="utf8") as f:
        temp_data_list = f.readlines()

    data_list = [data.strip().split(",") for data in temp_data_list[3:]]
    date_pos = 0
    year_set = set()
    data_dict: dict[dt, list[str]] = {}
    for data in data_list:
        dt_val = dt.strptime(data[date_pos], r"%Y/%m/%d")
        year_set.add(dt_val.strftime("%Y"))
        data_dict[dt_val] = data[date_pos + 1 :]
    return data_dict, year_set


def Input2Excel(file_path: Path, data_dict, year_set):
    wb = openpyxl.load_workbook(file_path)
    read_only_wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_list = read_only_wb.sheetnames
    ws_header_cnt = 2
    ws: Worksheet
    read_only_ws: ReadOnlyWorksheet
    for year in year_set:
        if year not in sheet_list:
            input(
                f"{year}のシートがありません。以下の手順でシートを作成してください。\n"
                f"・他のシートをコピーする\n"
                f"・シート名を{year}に変更する\n"
                f"・A1セルに{year}を入力する\n"
                f"・体調データが入力されている場合はすべて削除する\n"
                f"エンターを押すと終了します\n"
            )
            sys.exit()
        ws = wb[year]
        read_only_ws = read_only_wb[year]

        for dt_cell, condition_cell, comment_cell in zip(
            read_only_ws["A:A"][ws_header_cnt:],
            ws["C:C"][ws_header_cnt:],
            ws["D:D"][ws_header_cnt:],
        ):
            if data_dict.get(dt_cell.value) is None:
                continue

            data_list = data_dict[dt_cell.value]
            condition_str = data_list[0]
            condition_cell.value = (
                int(condition_str) if condition_str.isdigit() else condition_str
            )
            if len(data_list) > 1:
                comment_cell.value = data_list[1]
    wb.save(file_path)
    wb.close()
    read_only_wb.close()


def main(csv_path: str | None, xlsx_path: str | None):
    if not re.search(r"(iPhone|iPad)", platform.platform()):
        from FilePicker import GetFilePathByGUI

    if not csv_path:
        csv_path = GetFilePathByGUI(
            file_type=(["csvファイル", "*.csv"],),
        )[0]

    if not xlsx_path:
        xlsx_path = GetFilePathByGUI(
            file_type=(["xlsxファイル", "*.xlsx"],),
        )[0]

    condition_data_dict, year_set = ReadConditionCSV(Path(csv_path))
    Input2Excel(Path(xlsx_path), condition_data_dict, year_set)
    print("\n記録の入力が完了しました。")


if __name__ == "__main__":
    fire.Fire(main)
